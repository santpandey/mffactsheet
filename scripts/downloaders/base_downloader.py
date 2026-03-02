"""
Base Downloader - Shared logic for all mutual fund downloaders.

Each fund-specific downloader inherits from BaseFundDownloader and only
needs to implement `find_download_link()` with its own navigation logic.
"""

import sys
import time
import logging
import subprocess
from abc import ABC, abstractmethod
from pathlib import Path
from datetime import datetime
from typing import Optional, Tuple

import requests
from bs4 import BeautifulSoup
import openpyxl

# Shared constants
MAX_RETRIES = 3
RETRY_DELAY = 5       # seconds between retries
MIN_FILE_SIZE = 50 * 1024  # 50 KB
MIN_HOLDINGS = 30     # minimum rows to consider a valid file
MAX_PAGINATION = 10   # max pages to search

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

LOG_DIR = Path(__file__).parent.parent.parent / "logs"
LOG_DIR.mkdir(exist_ok=True)


def get_target_month() -> Tuple[int, int]:
    """Return (year, month) for the previous calendar month."""
    today = datetime.now()
    if today.month == 1:
        return today.year - 1, 12
    return today.year, today.month - 1


def make_session() -> requests.Session:
    """Create a requests session with browser-like headers."""
    session = requests.Session()
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Cache-Control": "max-age=0",
    })
    return session


def make_absolute(href: str, base_domain: str) -> str:
    """Convert a relative href to an absolute URL."""
    if href.startswith("http"):
        return href
    if href.startswith("//"):
        return "https:" + href
    if href.startswith("/"):
        return base_domain.rstrip("/") + href
    return base_domain.rstrip("/") + "/" + href


def fetch_page(session: requests.Session, url: str, logger: logging.Logger) -> Optional[BeautifulSoup]:
    """Fetch a URL and return a BeautifulSoup object, or None on failure."""
    try:
        response = session.get(url, timeout=30)
        response.raise_for_status()
        logger.info(f"  [OK] {url} ({response.status_code}, {len(response.content):,} bytes)")
        return BeautifulSoup(response.content, "html.parser")
    except requests.exceptions.RequestException as e:
        logger.error(f"  [ERR] Failed to fetch {url}: {e}")
        return None


def excel_links_from_soup(soup: BeautifulSoup) -> list:
    """Return all <a> tags whose href ends with .xlsx or .xls."""
    return [
        a for a in soup.find_all("a", href=True)
        if a["href"].lower().endswith((".xlsx", ".xls"))
    ]


def download_file(
    session: requests.Session,
    url: str,
    output_path: Path,
    logger: logging.Logger,
) -> bool:
    """Download a file with retry logic. Returns True on success."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            logger.info(f"Downloading (attempt {attempt}/{MAX_RETRIES}): {url}")
            response = session.get(url, timeout=60, stream=True)
            response.raise_for_status()

            content_type = response.headers.get("Content-Type", "")
            if not any(t in content_type.lower() for t in ("excel", "spreadsheet", "octet-stream", "zip")):
                logger.warning(f"Unexpected Content-Type: {content_type}")

            temp_path = output_path.with_suffix(".tmp")
            with open(temp_path, "wb") as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)

            size = temp_path.stat().st_size
            if size < MIN_FILE_SIZE:
                logger.error(f"File too small ({size} bytes) — likely not a valid Excel file")
                temp_path.unlink(missing_ok=True)
                return False

            temp_path.rename(output_path)
            logger.info(f"Saved {size:,} bytes -> {output_path.name}")
            return True

        except requests.exceptions.RequestException as e:
            logger.error(f"Download attempt {attempt} failed: {e}")
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY * attempt)

    return False


def validate_excel_generic(
    file_path: Path,
    fund_name_keywords: list,
    logger: logging.Logger,
) -> bool:
    """
    Generic Excel validation:
    - File opens without error
    - At least one sheet has MIN_HOLDINGS non-empty rows
    - At least one of fund_name_keywords appears in the first 10 rows (case-insensitive)
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        logger.info(f"Sheets found: {wb.sheetnames}")

        # Check fund name in any sheet's first 10 rows
        name_found = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(max_row=10, values_only=True):
                row_text = " ".join(str(c) for c in row if c).lower()
                if any(kw.lower() in row_text for kw in fund_name_keywords):
                    logger.info(f"Fund name confirmed in sheet '{sheet_name}'")
                    name_found = True
                    break
            if name_found:
                break

        if not name_found:
            logger.warning(
                f"Fund name keywords {fund_name_keywords} not found in first 10 rows "
                f"— proceeding anyway (some AMCs omit fund name in file)"
            )

        # Check data volume
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data_rows = sum(
                1 for row in ws.iter_rows(min_row=5, max_row=300, values_only=True)
                if any(c for c in row)
            )
            if data_rows >= MIN_HOLDINGS:
                logger.info(f"[OK] Sheet '{sheet_name}' has {data_rows} data rows — validation passed")
                wb.close()
                return True

        logger.error(f"No sheet has >= {MIN_HOLDINGS} data rows")
        wb.close()
        return False

    except Exception as e:
        logger.error(f"Validation error: {e}")
        return False


def run_extraction(logger: logging.Logger) -> None:
    """Trigger extract_all_funds.py after a successful download."""
    extract_script = Path(__file__).parent.parent / "extract_all_funds.py"
    try:
        result = subprocess.run(
            [sys.executable, str(extract_script)],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent.parent,
        )
        if result.returncode == 0:
            logger.info("[OK] Data extraction completed")
        else:
            logger.warning("Extraction finished with warnings")
            if result.stderr:
                logger.warning(result.stderr[:500])
    except Exception as e:
        logger.warning(f"Could not auto-extract: {e}")


class BaseFundDownloader(ABC):
    """
    Abstract base class for all fund downloaders.

    Subclasses must define:
        FUND_KEY          - short identifier, e.g. "sbi"
        FUND_DISPLAY_NAME - human-readable name
        BASE_DOMAIN       - root domain for making absolute URLs
        DOWNLOAD_DIR      - Path where Excel files are saved
        FUND_NAME_KEYWORDS- list of strings to validate the downloaded file

    Subclasses must implement:
        find_download_link(session, year, month, page) -> Optional[str]
        get_output_filename(year, month) -> Path
    """

    FUND_KEY: str = ""
    FUND_DISPLAY_NAME: str = ""
    BASE_DOMAIN: str = ""
    DOWNLOAD_DIR: Path = Path("excel-data")
    FUND_NAME_KEYWORDS: list = []

    def __init__(self):
        log_file = LOG_DIR / f"{self.FUND_KEY}_download_{datetime.now().strftime('%Y%m')}.log"
        self.logger = logging.getLogger(self.FUND_KEY)
        if not self.logger.handlers:
            self.logger.setLevel(logging.INFO)
            fmt = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
            fh = logging.FileHandler(log_file, encoding="utf-8")
            fh.setFormatter(fmt)
            sh = logging.StreamHandler(sys.stdout)
            sh.setFormatter(fmt)
            self.logger.addHandler(fh)
            self.logger.addHandler(sh)

    # ------------------------------------------------------------------
    # Abstract methods — each fund implements its own navigation logic
    # ------------------------------------------------------------------

    @abstractmethod
    def find_download_link(
        self,
        session: requests.Session,
        year: int,
        month: int,
        page: int,
    ) -> Optional[str]:
        """
        Search page `page` for the Excel download link for this fund.
        Returns the absolute URL string, or None if not found.
        """

    @abstractmethod
    def get_output_filename(self, year: int, month: int) -> Path:
        """Return the full Path where the downloaded file should be saved."""

    # ------------------------------------------------------------------
    # Shared orchestration
    # ------------------------------------------------------------------

    def file_exists(self, year: int, month: int) -> bool:
        path = self.get_output_filename(year, month)
        if path.exists():
            self.logger.info(f"Already exists: {path.name}")
            return True
        return False

    def download(self, year: int, month: int, force: bool = False) -> bool:
        """
        Full download pipeline:
        1. Check if file already exists
        2. Paginate through pages to find download link
        3. Download file
        4. Validate file
        5. Trigger extraction
        """
        self.logger.info("=" * 70)
        self.logger.info(f"[{self.FUND_DISPLAY_NAME}] Downloading {MONTH_NAMES[month-1]} {year}")
        self.logger.info("=" * 70)

        if not force and self.file_exists(year, month):
            self.logger.info("Skipping — file already downloaded.")
            return True

        self.DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
        session = make_session()

        # Paginate to find the link
        download_url = None
        for page in range(1, MAX_PAGINATION + 1):
            self.logger.info(f"Searching page {page}/{MAX_PAGINATION}...")
            download_url = self.find_download_link(session, year, month, page)
            if download_url:
                break
            time.sleep(1)

        if not download_url:
            self.logger.error(f"Download link not found after {MAX_PAGINATION} pages")
            return False

        output_path = self.get_output_filename(year, month)
        if not download_file(session, download_url, output_path, self.logger):
            self.logger.error("Download failed")
            return False

        if not validate_excel_generic(output_path, self.FUND_NAME_KEYWORDS, self.logger):
            self.logger.error("Validation failed — removing file")
            output_path.unlink(missing_ok=True)
            return False

        self.logger.info(f"[SUCCESS] {output_path.name}")
        run_extraction(self.logger)
        return True
