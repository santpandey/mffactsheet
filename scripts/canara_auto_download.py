"""
Automated Canara Robeco Monthly Portfolio Downloader
Downloads portfolio files from Canara Robeco website with validation and safeguards
"""

import os
import sys
import subprocess
import time
import logging
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional, Tuple
import requests
from bs4 import BeautifulSoup
import openpyxl

# Configuration
BASE_URL = "https://www.canararobeco.com/documents/statutory-disclosures/scheme-dashboard/scheme-monthly-portfolio/"
FUND_NAME_PATTERN = "Canara Robeco Large and Mid Cap Fund"
DOWNLOAD_DIR = Path(__file__).parent.parent / "excel-data" / "canara-robeco"
LOG_DIR = Path(__file__).parent.parent / "logs"
MAX_PAGINATION = 10
MAX_RETRIES = 3
RETRY_DELAY = 5  # seconds
MIN_FILE_SIZE = 50 * 1024  # 50 KB
MIN_HOLDINGS = 50

# Setup logging
LOG_DIR.mkdir(exist_ok=True)
log_file = LOG_DIR / f"canara_download_{datetime.now().strftime('%Y%m')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


def get_target_month() -> Tuple[int, int]:
    """
    Calculate the target month to download (previous month).
    Run between 5th-10th of current month to download previous month's data.
    
    Returns:
        Tuple of (year, month)
    """
    today = datetime.now()
    
    # Get previous month
    if today.month == 1:
        target_year = today.year - 1
        target_month = 12
    else:
        target_year = today.year
        target_month = today.month - 1
    
    return target_year, target_month


def get_output_filename(year: int, month: int) -> Path:
    """Generate output filename for downloaded file."""
    month_names = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    month_name = month_names[month - 1]
    filename = f"EQ-–-Canara-Robeco-Large-and-Mid-Cap-Fund-–-{month_name}-{year}.xlsx"
    return DOWNLOAD_DIR / filename


def file_already_exists(year: int, month: int) -> bool:
    """Check if file for the given month already exists."""
    output_file = get_output_filename(year, month)
    if output_file.exists():
        logger.info(f"File already exists: {output_file.name}")
        return True
    return False


def create_session() -> requests.Session:
    """Create a requests session with proper headers to avoid bot detection."""
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'none',
        'Cache-Control': 'max-age=0'
    })
    return session


def find_download_link(session: requests.Session, year: int, month: int, pagination: int) -> Optional[str]:
    """
    Search for the download link on a specific pagination page.
    
    Args:
        session: Requests session
        year: Target year
        month: Target month (1-12)
        pagination: Page number to search
        
    Returns:
        Download URL if found, None otherwise
    """
    url = f"{BASE_URL}?filteryear={year}&filtermonth={month:02d}&pagination={pagination}"
    
    try:
        logger.info(f"Searching pagination {pagination}: {url}")
        response = session.get(url, timeout=30)
        response.raise_for_status()
        
        logger.info(f"  [OK] Page loaded successfully (status: {response.status_code}, size: {len(response.content)} bytes)")
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Look for links containing the fund name
        # Common patterns: <a href="...xlsx">Fund Name</a> or data-download attributes
        links = soup.find_all('a', href=True)
        
        logger.info(f"  Found {len(links)} total links on page {pagination}")
        
        # Log all Excel/XLS links found
        excel_links = [link for link in links if link.get('href', '').endswith(('.xlsx', '.xls'))]
        logger.info(f"  Found {len(excel_links)} Excel file links")
        
        # Log fund-related links
        fund_related = []
        for link in links:
            href = link.get('href', '')
            text = link.get_text(strip=True)
            
            # Check for any "canara" or "large" or "mid cap" mentions
            if any(keyword in text.lower() for keyword in ['canara', 'large', 'mid cap', 'midcap']):
                fund_related.append(f"    - Text: '{text[:60]}...' | Href: {href[:80]}...")
        
        if fund_related:
            logger.info(f"  Found {len(fund_related)} fund-related links:")
            for item in fund_related[:10]:  # Log first 10
                logger.info(item)
        else:
            logger.info("  No fund-related links found on this page")
        
        for link in links:
            href = link.get('href', '')
            text = link.get_text(strip=True)
            
            # Check if this link is for our fund
            if FUND_NAME_PATTERN.lower() in text.lower() or FUND_NAME_PATTERN.lower() in href.lower():
                if href.endswith('.xlsx') or href.endswith('.xls'):
                    # Make absolute URL if relative
                    if href.startswith('http'):
                        download_url = href
                    elif href.startswith('/'):
                        download_url = f"https://www.canararobeco.com{href}"
                    else:
                        download_url = f"https://www.canararobeco.com/{href}"
                    
                    logger.info(f"  [MATCH] Found exact match: {download_url}")
                    return download_url
        
        # Try relaxed search: "large" AND "mid" OR "emerging equities"
        # Note: Fund was renamed from "Emerging Equities" to "Large and Mid Cap Fund"
        logger.info("  Trying relaxed search: 'large' AND 'mid' OR 'emerging equities'")
        for link in excel_links:
            text = link.get_text(strip=True).lower()
            href = link.get('href', '').lower()
            combined = text + ' ' + href
            
            # Match either the new name or the old name
            is_match = False
            if 'large' in combined and ('mid' in combined or 'midcap' in combined):
                is_match = True
            elif 'emerging' in combined and 'equities' in combined:
                is_match = True
            
            if is_match:
                href_actual = link.get('href', '')
                if href_actual.startswith('http'):
                    download_url = href_actual
                elif href_actual.startswith('/'):
                    download_url = f"https://www.canararobeco.com{href_actual}"
                else:
                    download_url = f"https://www.canararobeco.com/{href_actual}"
                
                logger.info(f"  [MATCH] Found via relaxed search: {download_url}")
                logger.info(f"         Link text: '{link.get_text(strip=True)[:100]}'")
                return download_url
        
        # Also check for download buttons with data attributes
        download_buttons = soup.find_all(['button', 'a'], attrs={'data-download': True})
        if download_buttons:
            logger.info(f"  Found {len(download_buttons)} download buttons")
            for button in download_buttons:
                if FUND_NAME_PATTERN.lower() in str(button).lower():
                    download_url = button.get('data-download') or button.get('href')
                    if download_url:
                        logger.info(f"  [MATCH] Found via button: {download_url}")
                        return download_url
        
        logger.info(f"  [NO MATCH] Target fund not found on page {pagination}")
        return None
        
    except requests.exceptions.RequestException as e:
        logger.error(f"  [ERROR] Failed to fetch page {pagination}: {e}")
        return None


def download_file(session: requests.Session, url: str, output_path: Path) -> bool:
    """
    Download file from URL with retry logic.
    
    Args:
        session: Requests session
        url: Download URL
        output_path: Where to save the file
        
    Returns:
        True if successful, False otherwise
    """
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            logger.info(f"Downloading (attempt {attempt}/{MAX_RETRIES}): {url}")
            
            response = session.get(url, timeout=60, stream=True)
            response.raise_for_status()
            
            # Check content type
            content_type = response.headers.get('Content-Type', '')
            if 'excel' not in content_type.lower() and 'spreadsheet' not in content_type.lower():
                logger.warning(f"Unexpected content type: {content_type}")
            
            # Download to temporary file first
            temp_path = output_path.with_suffix('.tmp')
            
            with open(temp_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            
            # Check file size
            file_size = temp_path.stat().st_size
            if file_size < MIN_FILE_SIZE:
                logger.error(f"Downloaded file too small: {file_size} bytes")
                temp_path.unlink()
                return False
            
            logger.info(f"Downloaded {file_size:,} bytes")
            
            # Rename to final name
            temp_path.rename(output_path)
            return True
            
        except requests.exceptions.RequestException as e:
            logger.error(f"Download attempt {attempt} failed: {e}")
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY * attempt)
            else:
                return False
    
    return False


def validate_excel_file(file_path: Path, year: int, month: int) -> bool:
    """
    Validate the downloaded Excel file.
    
    Checks:
    1. File can be opened by openpyxl
    2. Contains 'EQ' sheet
    3. Contains fund name
    4. Has sufficient holdings data
    
    Args:
        file_path: Path to Excel file
        year: Expected year
        month: Expected month
        
    Returns:
        True if valid, False otherwise
    """
    try:
        logger.info(f"Validating file: {file_path.name}")
        
        # Try to open with openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Check sheet names
        if 'EQ' not in wb.sheetnames:
            logger.error(f"Sheet 'EQ' not found. Available sheets: {wb.sheetnames}")
            return False
        
        ws = wb['EQ']
        
        # Check for fund name in first few rows
        # Note: The fund was previously named "Canara Robeco Emerging Equities" 
        # and was renamed to "Canara Robeco Large and Mid Cap Fund"
        # Accept both names for backward compatibility
        fund_name_found = False
        actual_fund_name = None
        for row in ws.iter_rows(max_row=10, values_only=True):
            row_text = ' '.join([str(cell) for cell in row if cell])
            if 'canara robeco' in row_text.lower():
                actual_fund_name = row_text.strip()
                # Accept either the new name or the old name (Emerging Equities)
                if ('large' in row_text.lower() and 'mid' in row_text.lower()) or \
                   ('emerging' in row_text.lower() and 'equities' in row_text.lower()):
                    fund_name_found = True
                    logger.info(f"  Fund name found: '{actual_fund_name}'")
                    break
        
        if not fund_name_found:
            if actual_fund_name:
                logger.error(f"Wrong fund! Expected 'Canara Robeco Large and Mid Cap Fund' (or 'Emerging Equities'), found: '{actual_fund_name}'")
            else:
                logger.error("Fund name not found in file")
            return False
        
        # Count data rows (rough estimate)
        data_rows = sum(1 for row in ws.iter_rows(min_row=5, max_row=200) if any(cell.value for cell in row))
        
        if data_rows < MIN_HOLDINGS:
            logger.error(f"Insufficient data rows: {data_rows} (minimum {MIN_HOLDINGS})")
            return False
        
        logger.info(f"[OK] File validation passed: ~{data_rows} data rows")
        wb.close()
        return True
        
    except Exception as e:
        logger.error(f"Validation error: {e}")
        return False


def download_monthly_portfolio(year: int, month: int, force: bool = False) -> bool:
    """
    Main function to download monthly portfolio.
    
    Args:
        year: Target year
        month: Target month (1-12)
        force: Force download even if file exists
        
    Returns:
        True if successful, False otherwise
    """
    logger.info("="*70)
    logger.info(f"Starting download for {year}-{month:02d}")
    logger.info("="*70)
    
    # Check if already exists
    if not force and file_already_exists(year, month):
        logger.info("File already exists. Use --force to re-download.")
        return True
    
    # Create output directory
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    
    # Create session
    session = create_session()
    
    # Search through pagination
    download_url = None
    for page in range(1, MAX_PAGINATION + 1):
        download_url = find_download_link(session, year, month, page)
        if download_url:
            break
        time.sleep(1)  # Be polite to the server
    
    if not download_url:
        logger.error(f"Could not find download link after checking {MAX_PAGINATION} pages")
        return False
    
    # Download file
    output_path = get_output_filename(year, month)
    if not download_file(session, download_url, output_path):
        logger.error("Download failed")
        return False
    
    # Validate file
    if not validate_excel_file(output_path, year, month):
        logger.error("File validation failed. Removing invalid file.")
        output_path.unlink()
        return False
    
    logger.info("="*70)
    logger.info(f"SUCCESS: Downloaded and validated: {output_path.name}")
    logger.info("="*70)
    
    # Auto-extract the downloaded file
    logger.info("")
    logger.info("Running extraction to generate JSON data...")
    try:
        import subprocess
        extract_script = Path(__file__).parent / "extract_all_funds.py"
        result = subprocess.run(
            [sys.executable, str(extract_script)],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent
        )
        if result.returncode == 0:
            logger.info("[OK] Data extraction completed successfully")
        else:
            logger.warning(f"Extraction completed with warnings. Check output above.")
    except Exception as e:
        logger.warning(f"Could not auto-extract data: {e}")
        logger.info("Run 'python scripts/extract_all_funds.py' manually to extract data")
    
    return True


def main():
    """Main entry point."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Download Canara Robeco monthly portfolio')
    parser.add_argument('--year', type=int, help='Target year (default: auto-detect)')
    parser.add_argument('--month', type=int, help='Target month 1-12 (default: auto-detect)')
    parser.add_argument('--force', action='store_true', help='Force download even if exists')
    parser.add_argument('--dry-run', action='store_true', help='Test without downloading')
    
    args = parser.parse_args()
    
    # Determine target month
    if args.year and args.month:
        year, month = args.year, args.month
    else:
        year, month = get_target_month()
    
    logger.info(f"Target: {year}-{month:02d}")
    
    if args.dry_run:
        logger.info("DRY RUN MODE - No files will be downloaded")
        output_file = get_output_filename(year, month)
        logger.info(f"Would download to: {output_file}")
        logger.info(f"File exists: {output_file.exists()}")
        return
    
    # Download
    success = download_monthly_portfolio(year, month, force=args.force)
    
    if success:
        logger.info("\n[SUCCESS] Download completed successfully")
        logger.info("Run extract_all_funds.py to process the new data")
        sys.exit(0)
    else:
        logger.error("\n[FAILED] Download failed")
        sys.exit(1)


if __name__ == "__main__":
    main()
